import openpyxl
import json

def parse_schedule_xlsx(xlsx_path):
    schedule_data = []
    try:
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active

        time_slots = []
        for cell in sheet[1]:
            if cell.value and str(cell.value).strip():
                time_slots.append(str(cell.value).strip())

        current_day = None
        for row_num in range(2, sheet.max_row + 1):
            row = sheet[row_num]
            first_cell_value = row[0].value

            if first_cell_value in ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]:
                current_day = str(first_cell_value).strip()
                continue

            if current_day:
                room = str(first_cell_value).strip() if first_cell_value else None
                if room:
                    for time_index, col_start in enumerate(range(1, sheet.max_column, 2)):
                        if time_index < len(time_slots):
                            time = time_slots[time_index]
                        else:
                            continue

                        course_cell = sheet.cell(row=row_num, column=col_start + 1)
                        teacher_cell = sheet.cell(row=row_num, column=col_start + 2)

                        course = str(course_cell.value).strip() if course_cell.value else None
                        teacher = str(teacher_cell.value).strip() if teacher_cell.value else None

                        if course:
                            parts = course.split('-')
                            course_name = parts[0] if parts else None
                            batch = parts[1] if len(parts) > 1 else None
                            section = parts[2] if len(parts) > 2 else None
                            teacher = teacher if teacher else None

                            class_info = {
                                "day": current_day,
                                "time": time,
                                "room": room,
                                "courseCode": course_name,
                                "teacherInitial": teacher,
                                "batch": batch,
                                "section": section
                            }
                            schedule_data.append(class_info)

    except FileNotFoundError:
        print(f"Error: XLSX file not found at {xlsx_path}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        return None
    return schedule_data

def write_json(schedule_array, output_path="routines.json"):
    template = {
        "semester": "Fall 2025",
        "department": "Software Engineering",
        "effectiveFrom": "__-__-2025",
        "schedule": schedule_array
    }

    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(template, f, indent=4, ensure_ascii=False)
        print(f"JSON file '{output_path}' created successfully!")
    except Exception as e:
        print(f"Error writing JSON to file: {e}")

def main():
    xlsx_path = r"schedule.xlsx"  # Or your path
    schedule_data = parse_schedule_xlsx(xlsx_path)
    if schedule_data:
        write_json(schedule_data)
    else:
        print("Failed to parse XLSX. JSON file not created.")

if __name__ == "__main__":
    main()